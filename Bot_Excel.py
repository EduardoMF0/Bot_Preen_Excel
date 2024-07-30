import pyautogui as auto
import time
from openpyxl import Workbook, load_workbook
import pyperclip
import tkinter as tk

# Caminho da planilha
file_path = r'Desktop\Planilha\Planilha_Excel.xlsx'

wb = load_workbook(file_path)
ws = wb.active


def store_number():
    global numero_inserido
    try:
        numero_inserido = int(entry.get())
    except ValueError:
        print("Por favor, insira um número válido.")
    root.destroy()

root = tk.Tk()
root.title("Entrada de Número")

tk.Label(root, text="Digite o número da linha da planilha: ").pack(pady=10)
entry = tk.Entry(root)
entry.pack(pady=10)

tk.Button(root, text="Executar", command=store_number).pack(pady=10)

numero_inserido = None

root.mainloop()

pos_linha_rk = numero_inserido
pos_linha_nome = numero_inserido
pos_linha_status = numero_inserido

time.sleep(1)

while True:

    os_lista = []
    aviso_pos = []

# --------------------------------- Funções ---------------------------------

    def get_rk(ws):
        global pos_linha_rk
        colu = 'B'
        pos_final = colu + str(pos_linha_rk)
        pos_final_value = ws[pos_final].value
        pos_linha_rk += 1
        return pos_final, pos_final_value


    def get_nome(ws):
        global pos_linha_nome
        colu = 'C'
        pos_final = colu + str(pos_linha_nome)
        pos_final_value = ws[pos_final].value
        pos_linha_nome += 1
        return pos_final_value


    def get_status(ws):
        global pos_linha_status
        colu = 'I'
        pos_final = colu+str(pos_linha_status)
        pos_final_value = ws[pos_final].value
        pos_linha_status += 1
        return pos_final, pos_final_value


    def buscar_os():
        try:
            os = auto.locateCenterOnScreen('oss.png', confidence=0.7)
            auto.click(os.x, os.y)
            if os:
                os_lista.append(os)
        except auto.ImageNotFoundException:
                os = None

        time.sleep(0.5)

        if os:
            down = auto.locateCenterOnScreen('download.png', confidence=0.7)
            auto.click(down.x, down.y)
            time.sleep(2.5)
            
            auto.press('enter')

            time.sleep(1)
        
        with auto.hold('alt'):
            auto.press('f4')

# --------------------------------- FIM - Funções ---------------------------------


    # INICIO DA EXECUÇÃO

    time.sleep(3)
    auto.scroll(2000)
    time.sleep(1)

    pesquisa = auto.locateCenterOnScreen('pesquisa.png', confidence=0.8)

    if pesquisa:
        auto.click(pesquisa.x, pesquisa.y)

    pos_rk, pos_rk_value = get_rk(ws)
    atual = get_nome(ws)
    pos_status, pos_status_value = get_status(ws)


    auto.press('backspace', presses=10)
    auto.write(pos_rk_value)
    time.sleep(0.5)
    auto.press('enter')

    # Confere Execução
    if pos_rk_value == None:
        break

    time.sleep(1.7)
    auto.click(x=401, y=378, clicks=3, interval=0.2) 
    with auto.hold('ctrl'):
        auto.press('c')

    conferir_nome = pyperclip.paste().strip().upper()

    # Confere dado atual
    if conferir_nome != atual:
        continue   
    
    auto.scroll(-560)
    time.sleep(0.6)

    # Confere elementos e clica nele

    conferir_imagem = 'lupa.png'

    imagem_posicao = list(auto.locateAllOnScreen(conferir_imagem, confidence=0.9))

    if imagem_posicao:
        for position in imagem_posicao:
            aviso_pos.append(position)
    else:
        print("Nenhum elemento encontrado.")

    elemen = len(aviso_pos)
    num = 0

    if aviso_pos:
        
        for elemen in imagem_posicao:
            center = aviso_pos[num]
            center_x = center.left + center.width / 2
            center_y = center.top + center.height / 2
            num += 1
            auto.click(center_x, center_y)
            time.sleep(1)
            buscar_os()
            time.sleep(1.2)


    time.sleep(0.6)
        
    auto.scroll(-700)

    time.sleep(1)

    imagem_posicao2 = list(auto.locateAllOnScreen(conferir_imagem, confidence=0.9))

    if imagem_posicao2:
            for position in imagem_posicao2:
                aviso_pos.append(position)
            else:
                print("Nenhum elemento encontrado.")


            for pos in imagem_posicao2:
                center = aviso_pos[num]
                center_x = center.left + center.width / 2
                center_y = center.top + center.height / 2
                num +=1
                auto.click(center_x, center_y)
                time.sleep(1)
                buscar_os()
                time.sleep(1.2)

    # Salva na Planilha

    if len(os_lista) > 1:
        ws[pos_status]= 'PASTA'
    else:
        ws[pos_status]= 'Documentos não disponivel na rede'
    
    wb.save(file_path)    

    time.sleep(0.7)

    # Resets

    aviso_pos.clear()
    os_lista.clear()
    num = 0
    elemen = 0